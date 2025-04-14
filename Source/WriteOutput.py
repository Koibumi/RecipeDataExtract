from RecipeProcess import *
from GroupProcess import *
from openpyxl import Workbook

def write_flattened_recipes(flat_rows: List[List[str]],
                     flattened_file: str = "flattened_recipes.xlsx") -> None:
    """
    将压平的 RecipeView 数据和倒排索引写入 XLSX 文件：
      - flattened_file: 每一行对应一个 RecipeView 对象（字段顺序为 workstations, conditions, result, ingredients, version）；
      - workstation_file: 每一行包含 workstation 名称和该名称在 flattened 表中的行索引（用分号分隔多个索引）；
      - ingredient_file: 每一行包含 ingredient 名称和对应的行索引（用分号分隔）。
    """
    # 写入压平后的 RecipeView 数据
    wb = Workbook()
    ws = wb.active
    ws.title = "data"
    ws.append(["workstations[]", "conditions[]", "result", "ingredients[]", "version"])
    for row in flat_rows:
        ws.append(row)
    wb.save(flattened_file)
    
    # 添加 metadata 工作表
    # 数据库要求的
    ws_meta = wb.create_sheet(title="fields")
    ws_meta.append(["name", "type", "title_zh", "title_en"])
    ws_meta.append(["workstations", "string", "工作站[]", "workstations[]"])
    ws_meta.append(["conditions", "string", "合成条件[]", "conditions[]"])
    ws_meta.append(["result", "string", "生成物", "result"])
    ws_meta.append(["ingredients", "string", "材料[]", "ingredients[]"])
    ws_meta.append(["version", "string", "版本", "version"])
    wb.save(flattened_file)
    
def write_flattened_groups(flat_rows: List[List[str]],
                     flattened_file: str = "flattened_groups.xlsx") -> None:
    """
    将压平的 GroupView 数据写入 XLSX 文件：
      - flattened_file: 每一行对应一个 GroupView 对象（字段顺序为 name, items[]）。
    """
    # 写入压平后的 GroupView 数据
    wb = Workbook()
    ws = wb.active
    ws.title = "data"
    ws.append(["name", "items[]"])
    for row in flat_rows:
        ws.append(row)
    wb.save(flattened_file)
    
    ws_meta = wb.create_sheet(title="fields")
    ws_meta.append(["name", "type", "title_zh", "title_en"])
    ws_meta.append(["name", "string", "组名称", "name"])
    ws_meta.append(["items", "string", "物品[]", "items[]"])
    wb.save(flattened_file)