from DataStructures import *
from openpyxl import load_workbook

def parse_recipe_row(row_data: dict) -> RecipeView:
    """
    从单行数据字典中恢复一个 RecipeView 对象。
    """
    # 处理 workstations 和 conditions，空字符串表示空集合
    ws = frozenset(row_data["workstations"].split(';')) if row_data["workstations"] else frozenset()
    cond = frozenset(row_data["conditions"].split(';')) if row_data["conditions"] else frozenset()

    # 处理 result 字段
    result = None
    if row_data["result"]:
        parts = row_data["result"].split(':')
        if len(parts) == 2:
            # 去掉可能的空白，并转换数量为整数
            result = ItemView(item=parts[0].strip(), quantity=int(parts[1].strip()))

    # 处理 ingredients 字段
    ingredients = frozenset()
    if row_data["ingredients"]:
        items = row_data["ingredients"].split(';')
        ing_list = []
        for item_str in items:
            if item_str:
                parts = item_str.split(':')
                if len(parts) == 2:
                    ing_list.append(ItemView(item=parts[0].strip(), quantity=int(parts[1].strip())))
        ingredients = frozenset(ing_list)

    version = row_data["version"]

    return RecipeView(
        workstations=ws,
        conditions=cond,
        result=result,
        ingredients=ingredients,
        version=version
    )

def load_recipes_from_xlsx(file_path: str) -> List[RecipeView]:
    """
    从 XLSX 文件中恢复 RecipeView 对象列表。
    
    XLSX 文件要求有如下列：
      - workstations: 多个工作台名称以分号分隔；
      - conditions: 多个条件以分号分隔；
      - result: 格式 "item:quantity"，如果为空则为 None；
      - ingredients: 多个 Ingredient 以分号分隔，每个格式 "item:quantity"；
      - version: 版本字符串。
    """
    recipes = []
    wb = load_workbook(file_path)
    sheet = wb['data']  # 读取 'data' 工作表

    # 获取表头
    headers = [cell.value for cell in sheet[1]]

    # 遍历每一行数据
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = dict(zip(headers, row))
        recipe = parse_recipe_row(row_data)
        recipes.append(recipe)
    return recipes